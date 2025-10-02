#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль для просмотра и работы с SQLite базой данных
Показывает все сохраненные пункты и информацию о БД
"""

from modules.database import DatabaseManager
import sqlite3
import os

def view_all_points():
    """Показывает все пункты из базы данных"""
    db_manager = DatabaseManager()
    
    print("📋 ВСЕ ПУНКТЫ ИЗ БАЗЫ ДАННЫХ")
    print("=" * 80)
    
    points = db_manager.get_all_points()
    
    if not points:
        print("❌ В базе данных нет пунктов")
        return
    
    current_file = None
    for point in points:
        filename, point_number, tag, seconds, content, created_at = point
        
        # Показываем имя файла только при смене
        if current_file != filename:
            current_file = filename
            print(f"\n📄 ФАЙЛ: {filename}")
            print("-" * 60)
        
        # Показываем пункт
        print(f"📌 Пункт {point_number}: {tag.title()} | {seconds} сек")
        print(f"📄 Текст: {content[:100]}{'...' if len(content) > 100 else ''}")
        print(f"⏰ Создан: {created_at}")
        print("-" * 40)

def view_database_info():
    """Показывает информацию о базе данных"""
    db_manager = DatabaseManager()
    
    print("🗄️ ИНФОРМАЦИЯ О БАЗЕ ДАННЫХ")
    print("=" * 50)
    
    try:
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            # Информация о таблицах
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = cursor.fetchall()
            
            print(f"📋 Таблицы в базе данных:")
            for table in tables:
                table_name = table[0]
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                count = cursor.fetchone()[0]
                print(f"   📊 {table_name}: {count} записей")
            
            # Размер базы данных
            db_size = os.path.getsize(db_manager.db_path)
            print(f"\n💾 Размер базы данных: {db_size} байт ({db_size/1024:.1f} КБ)")
            
    except sqlite3.Error as e:
        print(f"❌ Ошибка получения информации о БД: {e}")

def delete_database():
    """Удаляет базу данных"""
    db_manager = DatabaseManager()
    
    print("🗑️ УДАЛЕНИЕ БАЗЫ ДАННЫХ")
    print("=" * 40)
    
    try:
        if os.path.exists(db_manager.db_path):
            os.remove(db_manager.db_path)
            print(f"✅ База данных удалена: {db_manager.db_path}")
            print("💡 При следующем запуске анализа будет создана новая БД")
        else:
            print("ℹ️ База данных не найдена")
    except Exception as e:
        print(f"❌ Ошибка удаления БД: {e}")

def search_points_by_tag(tag):
    """Ищет пункты по тегу"""
    db_manager = DatabaseManager()
    
    print(f"🔍 ПОИСК ПУНКТОВ ПО ТЕГУ: {tag.upper()}")
    print("=" * 60)
    
    try:
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT 
                    f.filename,
                    p.point_number,
                    p.tag,
                    p.seconds,
                    p.content,
                    p.created_at
                FROM points p
                JOIN files f ON p.file_id = f.id
                WHERE LOWER(p.tag) LIKE LOWER(?)
                ORDER BY f.filename, p.point_number
            ''', (f'%{tag}%',))
            
            points = cursor.fetchall()
            
            if not points:
                print(f"❌ Пункты с тегом '{tag}' не найдены")
                return
            
            print(f"📊 Найдено пунктов: {len(points)}")
            print("-" * 60)
            
            current_file = None
            for point in points:
                filename, point_number, point_tag, seconds, content, created_at = point
                
                if current_file != filename:
                    current_file = filename
                    print(f"\n📄 ФАЙЛ: {filename}")
                    print("-" * 40)
                
                print(f"📌 Пункт {point_number}: {point_tag.title()} | {seconds} сек")
                print(f"📄 Текст: {content[:80]}{'...' if len(content) > 80 else ''}")
                print("-" * 30)
                
    except sqlite3.Error as e:
        print(f"❌ Ошибка поиска: {e}")

def get_points_summary():
    """Показывает краткую сводку по пунктам"""
    db_manager = DatabaseManager()
    
    try:
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            # Общее количество пунктов
            cursor.execute("SELECT COUNT(*) FROM points")
            total_points = cursor.fetchone()[0]
            
            # Количество файлов
            cursor.execute("SELECT COUNT(*) FROM files")
            total_files = cursor.fetchone()[0]
            
            # Статистика по тегам в разбивке по папкам
            cursor.execute("""
                SELECT 
                    f.file_path,
                    p.tag, 
                    COUNT(*) as count
                FROM points p
                JOIN files f ON p.file_id = f.id
                GROUP BY f.file_path, p.tag 
                ORDER BY f.file_path, p.tag
            """)
            folder_tag_stats = cursor.fetchall()
            
            print(f"📁 Файлов: {total_files}")
            print(f"📋 Пунктов: {total_points}")
            print("🏷️ Статистика по папкам и тегам:")
            
            # Группируем по папкам
            folder_data = {}
            for file_path, tag, count in folder_tag_stats:
                # Извлекаем папку из пути
                from pathlib import Path
                path_parts = Path(file_path).parts
                if len(path_parts) >= 2:
                    folder_name = path_parts[-2]  # Предпоследняя часть пути
                else:
                    folder_name = "Корень"
                
                if folder_name not in folder_data:
                    folder_data[folder_name] = {}
                
                if tag not in folder_data[folder_name]:
                    folder_data[folder_name][tag] = 0
                
                folder_data[folder_name][tag] += count
            
            for folder_name, tags in folder_data.items():
                if folder_name != "Корень":
                    print(f"  📁 {folder_name}:")
                    for tag, count in tags.items():
                        print(f"    📌 {tag}: {count} пунктов")
            
            # Проверяем пункты короче 30 секунд ТОЛЬКО для тега 'губер'
            cursor.execute("""
                SELECT 
                    f.file_path,
                    p.tag, 
                    COUNT(*) as count
                FROM points p
                JOIN files f ON p.file_id = f.id
                WHERE p.seconds < 30 AND p.tag = 'губер'
                GROUP BY f.file_path, p.tag 
                ORDER BY f.file_path, p.tag
            """)
            short_points_stats = cursor.fetchall()
            
            if short_points_stats:
                print("⚠️ Пункты 'губер' короче 30 сек (исключены из отчетов):")
                
                # Группируем короткие пункты по папкам
                short_folder_data = {}
                for file_path, tag, count in short_points_stats:
                    path_parts = Path(file_path).parts
                    if len(path_parts) >= 2:
                        folder_name = path_parts[-2]
                    else:
                        folder_name = "Корень"
                    
                    if folder_name not in short_folder_data:
                        short_folder_data[folder_name] = {}
                    
                    if tag not in short_folder_data[folder_name]:
                        short_folder_data[folder_name][tag] = 0
                    
                    short_folder_data[folder_name][tag] += count
                
                for folder_name, tags in short_folder_data.items():
                    if folder_name != "Корень":
                        print(f"  📁 {folder_name}:")
                        for tag, count in tags.items():
                            print(f"    ⏱️ {tag}: {count} пунктов")
            
            # Проверяем, есть ли пункты без короткого описания
            cursor.execute("""
                SELECT COUNT(*) FROM points 
                WHERE (short_content IS NULL OR short_content = '' OR LENGTH(TRIM(short_content)) = 0)
            """)
            points_without_short = cursor.fetchone()[0]
            
            print("=" * 40)
            if points_without_short > 0:
                print(f"🤖 Формируем краткое описание для {points_without_short} пунктов...")
                process_points_with_gpt(db_manager)
            else:
                print("✅ Все пункты имеют краткое описание")
            
            print("=" * 40)
            print("📊 Создание Excel отчета...")
            create_excel_report()
            
    except sqlite3.Error as e:
        print(f"❌ Ошибка получения сводки: {e}")


def create_excel_report():
    """Создает Excel отчет с группировкой по папкам и тегам - каждый тег в отдельном файле, используя шаблон"""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import re
        import yaml
        from pathlib import Path
        import os
        
        print("📊 Создание Excel отчета")
        print("=" * 40)
        
        # Проверяем наличие Pillow для работы с изображениями
        try:
            import PIL
            pillow_available = True
        except ImportError:
            pillow_available = False
            print("⚠️ Предупреждение: Pillow не установлен, логотипы не будут добавлены")
        
        # Загружаем конфигурацию
        try:
            with open('config.yaml', 'r', encoding='utf-8') as file:
                config = yaml.safe_load(file)
                reports_config = config.get('reports', {})
                export_path = reports_config.get('path', 'reports')
                year = reports_config.get('year', '2024')
                min_duration = reports_config.get('min_duration_seconds', 30)
                docs_config = reports_config.get('docs', {})
        except:
            export_path = 'reports'
            year = '2024'
            min_duration = 30
            docs_config = {}
        
        # Создаем папку для экспорта, если её нет
        export_dir = Path(export_path)
        export_dir.mkdir(exist_ok=True)
        
        print(f"📁 Папка экспорта: {export_dir.absolute()}")
        print(f"📅 Год: {year}")
        print(f"⏱️ Минимальная длительность: {min_duration} сек (только для тега 'губер')")
        
        # Путь к шаблону
        template_path = Path("../static/template.xlsx")
        if not template_path.exists():
            print(f"❌ Шаблон не найден: {template_path}")
            return False
        
        print(f"📋 Используем шаблон: {template_path}")
        
        db_manager = DatabaseManager()
        
        # Получаем все файлы с их путями
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            # Получаем файлы с группировкой по папкам
            cursor.execute("""
                SELECT 
                    filename,
                    file_path,
                    file_type,
                    encoding,
                    created_at
                FROM files 
                ORDER BY file_path
            """)
            
            files = cursor.fetchall()
            
            if not files:
                print("❌ В базе данных нет файлов для отчета")
                return False
            
            # Группируем файлы по папкам
            folder_groups = {}
            for file in files:
                filename, file_path, file_type, encoding, created_at = file
                
                # Извлекаем папку из пути (например, ИЮЛЬ/КП/ -> КП)
                path_parts = Path(file_path).parts
                if len(path_parts) >= 2:
                    folder_name = path_parts[-2]  # Берем предпоследнюю часть пути
                else:
                    folder_name = "Корень"
                
                if folder_name not in folder_groups:
                    folder_groups[folder_name] = []
                
                folder_groups[folder_name].append({
                    'filename': filename,
                    'file_path': file_path,
                    'file_type': file_type,
                    'encoding': encoding,
                    'created_at': created_at
                })
            
            # Создаем папку и Excel файлы для каждой папки
            for folder_name, folder_files in folder_groups.items():
                if folder_name == "Корень":
                    continue  # Пропускаем файлы в корне
                
                print(f"\n📁 Обрабатываем папку: {folder_name}")
                
                # Создаем папку для этой группы
                folder_dir = export_dir / (folder_name + datetime.now().strftime("_%Y%m%d_%H%M%S"))
                folder_dir.mkdir(exist_ok=True)
                
                # Получаем конфигурацию для этой папки
                folder_config = docs_config.get(folder_name, {})
                month = folder_config.get('month', 'месяц')
                full_name = folder_config.get('full_name', 'Ф.И.О.')
                company_short = folder_config.get('company_short', 'CXXX')
                company_full = folder_config.get('company_full', 'CFXXX')
                logo_path = folder_config.get('logo_path', '')
                
                # Получаем все теги для этой папки
                cursor.execute("""
                    SELECT DISTINCT p.tag
                    FROM points p
                    JOIN files f ON p.file_id = f.id
                    WHERE f.file_path LIKE ?
                    ORDER BY p.tag
                """, (f'%{folder_name}%',))
                
                tags = cursor.fetchall()
                
                # Создаем Excel файл для каждого тега
                for tag_row in tags:
                    tag = tag_row[0]
                    print(f"   🏷️ Создаем файл для тега: {tag}")
                    
                    # Создаем Excel файл на основе шаблона
                    excel_filename = folder_dir / f"{tag}.xlsx"
                    
                    # Загружаем шаблон
                    wb = load_workbook(template_path)
                    ws = wb.active
                    
                    # Добавляем логотип в верхний левый угол
                    add_logo_to_report(ws, logo_path)
                    
                    # Заполняем шапку отчета
                    fill_report_header(ws, year, month, full_name, company_short, company_full)
                    
                    # Получаем пункты для этого тега в этой папке (фильтруем по минимальной длительности ТОЛЬКО для 'губер')
                    cursor.execute("""
                        SELECT 
                            p.point_number,
                            f.filename,
                            p.tag,
                            p.seconds,
                            p.content,
                            p.short_content,
                            p.created_at
                        FROM points p
                        JOIN files f ON p.file_id = f.id
                        WHERE f.file_path LIKE ? AND p.tag = ? AND (p.tag != 'губер' OR p.seconds >= ?)
                        ORDER BY f.filename, p.point_number
                    """, (f'%{folder_name}%', tag, min_duration))
                    
                    points = cursor.fetchall()
                    
                    # Сортируем по дате/времени из имени файла (от ранней к поздней)
                    if points:
                        from datetime import datetime
                        def _dt_key(rec):
                            # rec: (point_number, filename, tag, seconds, content, short_content, created_at)
                            fn = rec[1]
                            m = re.search(r'(\d{1,2})\.(\d{2})\s+на\s+(\d{1,2})-(\d{2})\D?', fn)
                            if m:
                                d, M, h, m_ = m.groups()
                                try:
                                    return datetime(int(year), int(M), int(d), int(h), int(m_))
                                except:
                                    return datetime.min
                            m2 = re.search(r'(\d{1,2})\.(\d{2})', fn)
                            if m2:
                                d, M = m2.groups()
                                try:
                                    return datetime(int(year), int(M), int(d))
                                except:
                                    return datetime.min
                            return datetime.min
                        points = sorted(points, key=_dt_key)
                    
                    if points:
                        # Начинаем запись данных с 15-й строки (шапка в 14-й)
                        start_row = 15
                        
                        # Создаем бордер для ячеек
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Данные
                        row = start_row
                        total_seconds = 0
                        for i, point in enumerate(points, 1):
                            point_number, filename, tag, seconds, content, short_content, created_at = point
                            
                            # Извлекаем дату и время из имени файла (поддержка 1-2 цифр дня/часа и возможной буквы после времени)
                            dt_match = re.search(r'(\d{1,2})\.(\d{2})\s+на\s+(\d{1,2})-(\d{2})\D?', filename)
                            if dt_match:
                                d, M, h, m_ = dt_match.groups()
                                date_part = f"{int(d):02d}.{M}"
                                time_part = f"{int(h):02d}:{m_}"
                                broadcast_date_time = f"{date_part}.{year} {time_part}"
                            else:
                                d_match = re.search(r'(\d{1,2})\.(\d{2})', filename)
                                if d_match:
                                    d, M = d_match.groups()
                                    date_part = f"{int(d):02d}.{M}"
                                    broadcast_date_time = f"{date_part}.{year}"
                                else:
                                    broadcast_date_time = "Дата не найдена"
                            
                            # Используем сокращенный текст, если есть, иначе полный
                            display_text = short_content if short_content else content[:100] + "..." if len(content) > 100 else content
                            
                            # Заполняем данные начиная со второй колонки (B):
                            # A | B | C | D | E
                            #   | № п/п | Дата и время выхода в эфир | Тема информационного материала | Хронометраж (мин/сек.)
                            ws.cell(row=row, column=2, value=i)  # № п/п (колонка B)
                            ws.cell(row=row, column=3, value=broadcast_date_time)  # Дата и время выхода в эфир (колонка C)
                            ws.cell(row=row, column=4, value=display_text)  # Тема информационного материала (колонка D)
                            
                            # Оставляем секунды как есть
                            ws.cell(row=row, column=5, value=seconds)  # Хронометраж в секундах (колонка E)
                            
                            # Добавляем бордер к каждой ячейке строки данных (колонки B-E)
                            for col in range(2, 6):  # Колонки B, C, D, E
                                cell = ws.cell(row=row, column=col)
                                cell.border = thin_border
                            
                            # Суммируем общее время
                            total_seconds += seconds
                            
                            row += 1
                        
                        # Добавляем итоги в конце без отступа
                        total_materials = len(points)
                        total_minutes = total_seconds // 60
                        total_remaining_seconds = total_seconds % 60
                        total_time = f"{total_minutes} мин {total_remaining_seconds:02d} сек"
                        
                        # Итоговая строка (без отступа от данных)
                        summary_row = row  # Сразу после последней строки данных
                        ws.cell(row=summary_row, column=3, value="Итого:")  # Начинаем с третьего столбца (C)
                        ws.cell(row=summary_row, column=4, value=f"{total_materials} информационных материала")
                        ws.cell(row=summary_row, column=5, value=total_time)
                        
                        # Добавляем бордер к итоговой строке (колонки B-E)
                        for col in range(2, 6):  # Колонки B, C, D, E (со второго столбца)
                            cell = ws.cell(row=summary_row, column=col)
                            cell.border = thin_border
                        
                        # НЕ изменяем ширину колонок - сохраняем из шаблона
                    
                    # Сохраняем Excel файл
                    wb.save(excel_filename)
                    print(f"      ✅ Excel файл создан: {excel_filename.name}")
            
            print(f"\n🎉 Создание Excel отчетов завершено!")
            print(f"📁 Обработано папок: {len(folder_groups)}")
            
            return True
            
    except ImportError as e:
        print(f"❌ Не установлены необходимые библиотеки: {e}")
        print("💡 Установите: pip install pandas openpyxl")
        return False
    except Exception as e:
        print(f"❌ Ошибка создания Excel отчета: {e}")
        return False

def add_logo_to_report(ws, logo_path):
    """Добавляет логотип в верхний левый угол отчета с фиксированной шириной 150px"""
    try:
        from pathlib import Path
        
        if not logo_path or not Path(logo_path).exists():
            print(f"      ⚠️ Логотип не найден или путь пустой: {logo_path}")
            return
        
        try:
            from openpyxl.drawing.image import Image
            
            # Создаем объект изображения
            img = Image(logo_path)
            
            # Получаем оригинальные размеры
            original_width = img.width
            original_height = img.height
            
            # Вычисляем коэффициент масштабирования для ширины 150px
            scale_factor = 150 / original_width
            
            # Устанавливаем фиксированную ширину 150px и пропорциональную высоту
            img.width = 150
            img.height = int(original_height * scale_factor)

            # Добавляем логотип в ячейку A1 (верхний левый угол)
            ws.add_image(img, 'A1')
            
        except ImportError as e:
            if "Pillow" in str(e):
                print(f"      ⚠️ Для добавления логотипов необходимо установить Pillow:")
                print(f"      💡 Выполните: pip install Pillow")
                print(f"      📋 Отчет будет создан без логотипа")
            else:
                print(f"      ⚠️ Ошибка импорта изображения: {e}")
        
    except Exception as e:
        print(f"      ⚠️ Ошибка добавления логотипа: {e}")
        print(f"      📋 Отчет будет создан без логотипа")

def fill_report_header(ws, year, month, full_name, company_short, company_full):
    """Заполняет шапку отчета данными из конфигурации"""
    try:
        # Ищем и заменяем плейсхолдеры в шапке
        for row in range(1, 20):  # Проверяем первые 20 строк
            for col in range(1, 10):  # Проверяем первые 10 колонок
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    # Заменяем плейсхолдеры
                    cell.value = cell.value.replace('YXXX', year)
                    cell.value = cell.value.replace('MXXX', month)
                    cell.value = cell.value.replace('GXXX', full_name)
                    cell.value = cell.value.replace('CXXX', company_short)
                    cell.value = cell.value.replace('CFXXX', company_full)
    except Exception as e:
        print(f"⚠️ Ошибка заполнения шапки: {e}")


def create_simple_excel_report():
    """Создает простой Excel отчет по папкам без вкладок - все данные в одной таблице"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import re
        import yaml
        from pathlib import Path
        
        print("📊 СОЗДАНИЕ ПРОСТОГО EXCEL ОТЧЕТА")
        print("=" * 40)
        
        # Загружаем конфигурацию
        try:
            with open('config.yaml', 'r', encoding='utf-8') as file:
                config = yaml.safe_load(file)
                reports_config = config.get('reports', {})
                export_path = reports_config.get('path', 'reports')
                min_duration = reports_config.get('min_duration_seconds', 30)
        except:
            export_path = 'reports'
            min_duration = 30
        
        # Создаем папку для экспорта, если её нет
        export_dir = Path(export_path)
        export_dir.mkdir(exist_ok=True)
        
        print(f"📁 Папка для экспорта: {export_dir.absolute()}")
        print(f"⏱️ Минимальная длительность: {min_duration} сек (только для тега 'губер')")
        
        db_manager = DatabaseManager()
        
        # Получаем все файлы с их путями
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            # Получаем файлы с группировкой по папкам
            cursor.execute("""
                SELECT 
                    filename,
                    file_path,
                    file_type,
                    encoding,
                    created_at
                FROM files 
                ORDER BY file_path
            """)
            
            files = cursor.fetchall()
            
            if not files:
                print("❌ В базе данных нет файлов для отчета")
                return False
            
            # Группируем файлы по папкам
            folder_groups = {}
            for file in files:
                filename, file_path, file_type, encoding, created_at = file
                
                # Извлекаем папку из пути (например, ИЮЛЬ/КП/ -> КП)
                path_parts = Path(file_path).parts
                if len(path_parts) >= 2:
                    folder_name = path_parts[-2]  # Берем предпоследнюю часть пути
                else:
                    folder_name = "Корень"
                
                if folder_name not in folder_groups:
                    folder_groups[folder_name] = []
                
                folder_groups[folder_name].append({
                    'filename': filename,
                    'file_path': file_path,
                    'file_type': file_type,
                    'encoding': encoding,
                    'created_at': created_at
                })
            
            # Создаем Excel файл для каждой папки
            for folder_name, folder_files in folder_groups.items():
                if folder_name == "Корень":
                    continue  # Пропускаем файлы в корне
                
                print(f"\n📁 Обрабатываем папку: {folder_name}")
                
                # Создаем Excel файл
                excel_filename = export_dir / f"{folder_name}.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = folder_name
                
                # Получаем все пункты для этой папки (фильтруем по минимальной длительности ТОЛЬКО для 'губер')
                cursor.execute("""
                    SELECT 
                        p.point_number,
                        f.filename,
                        p.tag,
                        p.seconds,
                        p.content,
                        p.short_content,
                        p.created_at
                    FROM points p
                    JOIN files f ON p.file_id = f.id
                    WHERE f.file_path LIKE ? AND (p.tag != 'губер' OR p.seconds >= ?)
                    ORDER BY f.filename, p.point_number
                """, (f'%{folder_name}%', min_duration))
                
                points = cursor.fetchall()
                
                # Сортируем по дате/времени из имени файла (от ранней к поздней)
                if points:
                    from datetime import datetime
                    def _dt_key(rec):
                        # rec: (point_number, filename, tag, seconds, content, short_content, created_at)
                        fn = rec[1]
                        m = re.search(r'(\d{1,2})\.(\d{2})\s+на\s+(\d{1,2})-(\d{2})\D?', fn)
                        if m:
                            d, M, h, m_ = m.groups()
                            try:
                                # Для простого отчета год не задаётся явно; возьмем текущий год
                                return datetime(datetime.now().year, int(M), int(d), int(h), int(m_))
                            except:
                                return datetime.min
                        m2 = re.search(r'(\d{1,2})\.(\d{2})', fn)
                        if m2:
                            d, M = m2.groups()
                            try:
                                return datetime(datetime.now().year, int(M), int(d))
                            except:
                                return datetime.min
                        return datetime.min
                    points = sorted(points, key=_dt_key)
                
                if points:
                    # Заголовки
                    headers = ['№', 'Дата эфира', 'Тег', 'Текст (сокращенный)', 'Секунды']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Данные
                    row = 2
                    for i, point in enumerate(points, 1):
                        point_number, filename, tag, seconds, content, short_content, created_at = point
                        
                        # Извлекаем дату из имени файла (поддержка 1-2 цифр дня)
                        d_match = re.search(r'(\d{1,2})\.(\d{2})', filename)
                        if d_match:
                            d, M = d_match.groups()
                            broadcast_date = f"{int(d):02d}.{M}"
                        else:
                            broadcast_date = "Дата не найдена"
                        
                        # Используем сокращенный текст, если есть, иначе полный
                        display_text = short_content if short_content else content[:100] + "..." if len(content) > 100 else content
                        
                        ws.cell(row=row, column=1, value=i)  # Простой порядковый номер строки
                        ws.cell(row=row, column=2, value=broadcast_date)
                        ws.cell(row=row, column=3, value=tag.title())
                        ws.cell(row=row, column=4, value=display_text)
                        ws.cell(row=row, column=5, value=seconds)
                        
                        row += 1
                    
                    # Автоподбор ширины столбцов
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Сохраняем Excel файл
                wb.save(excel_filename)
                print(f"   ✅ Excel файл создан: {excel_filename.name}")
            
            print(f"\n🎉 Создание простых Excel отчетов завершено!")
            print(f"📁 Обработано папок: {len(folder_groups)}")
            
            return True
            
    except ImportError as e:
        print(f"❌ Не установлены необходимые библиотеки: {e}")
        print("💡 Установите: pip install pandas openpyxl")
        return False
    except Exception as e:
        print(f"❌ Ошибка создания Excel отчета: {e}")
        return False


def process_points_with_gpt(db_manager):
    """Обрабатывает пункты через GPT для создания краткого описания"""
    try:
        # Импортируем функцию GPT
        from modules.text_shortener import shorten_text
        
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            # Получаем пункты без краткого описания
            cursor.execute("""
                SELECT id, content, tag, point_number 
                FROM points 
                WHERE (short_content IS NULL OR short_content = '' OR LENGTH(TRIM(short_content)) = 0)
                ORDER BY id
            """)
            
            points_to_process = cursor.fetchall()
            
            if not points_to_process:
                print("✅ Все пункты уже обработаны")
                return
            

            
            processed = 0
            for point_id, content, tag, point_number in points_to_process:
                try:
                    # Отправляем на GPT
                    short_content = shorten_text(content)
                    
                    if short_content and not short_content.startswith("Ошибка"):
                        # Сохраняем результат
                        if db_manager.update_point_short_content(point_id, short_content):
                            print(f"✅ Пункт {point_number} ({tag}): {short_content}")
                            processed += 1
                        else:
                            print(f"❌ Ошибка сохранения пункта {point_number}")
                    else:
                        print(f"⚠️ Ошибка GPT для пункта {point_number}: {short_content}")
                    
                    # Небольшая пауза между запросами
                    import time
                    time.sleep(1)
                    
                except Exception as e:
                    print(f"❌ Ошибка обработки пункта {point_id}: {e}")
                    continue
            
            print(f"🎉 Обработка завершена: {processed}/{len(points_to_process)} пунктов")
            
    except ImportError:
        print("❌ Модуль text_shortener не найден")
    except Exception as e:
        print(f"❌ Ошибка GPT обработки: {e}")